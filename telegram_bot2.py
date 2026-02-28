#!/usr/bin/env python3
"""
Telegram Bot 2 — Ark/GLM-4 AI Chatbot + OpenClaw Skills + Debug Panel
──────────────────────────────────────────────────────────────────────
AI chatbot powered by Ark (BytePlus) GLM-4.
Forwards /skill commands to OpenClaw Gateway for skill execution.
Debug panel + utility skills for VPS management via Telegram.
"""
import os, subprocess, logging, asyncio, time, json, re, math, csv, io
from datetime import datetime, timedelta
from pathlib import Path
import httpx
from telegram_trend_bot_router import TrendBotRouter
from skill_menu_engine import SkillMenuEngine
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes
)

# === CONFIG ===================================================================
TELEGRAM_TOKEN = "8258288878:AAGCxzBuTQ8joXVUhDS_ogtYh4j_4Pjc0Jc"

# AI Model Providers
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_API_KEY = "sk-72e5d53c4468437687d5f3afe383b3b0"

ARK_API_URL = "https://ark.ap-southeast.bytepluses.com/api/v3/chat/completions"
ARK_GLM_KEY = "1f11c03a-b90f-4225-8084-d132994f0dfe"
ARK_KIMI_KEY = "aa2f56e1-3a25-4e32-b27b-dc0e7f8da1d1"

SUMOPOD_API_URL = "https://ai.sumopod.com/v1/chat/completions"
SUMOPOD_API_KEY = "sk-OEQrtl_iYyey-IVy5BTmfQ"

# YouTube Data API v3 (for FunnelTrendResearcher)
YOUTUBE_API_KEY = "AIzaSyCvwuzvg6K4aef3dAuqC6nC_uPY3Eh4fms"

# Available models: key -> (display_name, api_url, api_key, model_id)
AVAILABLE_MODELS = {
    "v3":     ("DeepSeek V3",     DEEPSEEK_API_URL, DEEPSEEK_API_KEY, "deepseek-chat"),
    "r1":     ("DeepSeek R1",     DEEPSEEK_API_URL, DEEPSEEK_API_KEY, "deepseek-reasoner"),
    "glm4":   ("GLM-4",          ARK_API_URL,      ARK_GLM_KEY,      "glm-4-7-251222"),
    "kimi":   ("Kimi K2",        ARK_API_URL,      ARK_KIMI_KEY,     "kimi-k2-250905"),
    "sumo":   ("SumoPod GPT-4.1", SUMOPOD_API_URL,  SUMOPOD_API_KEY,  "gpt-4.1-nano"),
    "codex":  ("SumoPod Codex",   SUMOPOD_API_URL,  SUMOPOD_API_KEY,  "gpt-5.1-codex-mini"),
}
DEFAULT_MODEL = "kimi"

# Per-user model selection
user_model_selection = {}

# OpenClaw Gateway (for skill execution)
OPENCLAW_HOST = "http://127.0.0.1:18789"
OPENCLAW_TOKEN = "6b833e046e3c2f36b9aae4c5134ee56bdf4a9a04dc1f054d"
OPENCLAW_MODEL = "sumopod/gpt-4.1-nano"

# Admin user IDs
ADMIN_IDS = {231743189}

# Notes file
NOTES_FILE = "/root/.bot2_notes.json"

SYSTEM_PROMPT = """Kamu adalah asisten AI yang ramah dan helpful. 
Jawab pertanyaan user dengan jelas, ringkas, dan dalam bahasa yang sama dengan user.
Jika user berbicara Bahasa Indonesia, jawab dalam Bahasa Indonesia.
Jika user berbicara English, jawab dalam English."""

logging.basicConfig(format="%(asctime)s [%(levelname)s] %(name)s: %(message)s", level=logging.INFO)
log = logging.getLogger("bot2")

# === CONVERSATION HISTORY =====================================================
conversation_history = {}
MAX_HISTORY = 20

def add_to_history(user_id, role, content):
    if user_id not in conversation_history:
        conversation_history[user_id] = []
    conversation_history[user_id].append({"role": role, "content": content})
    if len(conversation_history[user_id]) > MAX_HISTORY:
        conversation_history[user_id] = conversation_history[user_id][-MAX_HISTORY:]

def get_history(user_id):
    return conversation_history.get(user_id, [])

# === HELPER: Run shell command ================================================
def run_shell(cmd, timeout=30):
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)
        output = result.stdout + result.stderr
        if not output.strip():
            output = "(no output)"
        if len(output) > 3900:
            output = output[:3900] + "\n\n... (truncated)"
        return output
    except subprocess.TimeoutExpired:
        return "⏱ Command timed out after {}s".format(timeout)
    except Exception as e:
        return "❌ Error: {}".format(e)

def is_admin(user_id):
    return user_id in ADMIN_IDS

# === NOTES SYSTEM =============================================================
def load_notes():
    try:
        with open(NOTES_FILE, "r") as f:
            return json.load(f)
    except:
        return []

def save_notes(notes):
    with open(NOTES_FILE, "w") as f:
        json.dump(notes, f, indent=2, ensure_ascii=False)

# === OPENCLAW GATEWAY CLIENT ==================================================
async def ask_openclaw(message: str, timeout_sec=120) -> str:
    """Forward a message to OpenClaw Gateway for skill execution."""
    try:
        async with httpx.AsyncClient(timeout=timeout_sec) as client:
            response = await client.post(
                OPENCLAW_HOST + "/v1/chat/completions",
                headers={
                    "Authorization": "Bearer " + OPENCLAW_TOKEN,
                    "Content-Type": "application/json"
                },
                json={
                    "model": OPENCLAW_MODEL,
                    "messages": [
                        {"role": "user", "content": message}
                    ],
                    "max_tokens": 4096
                }
            )
            response.raise_for_status()
            data = response.json()
            reply = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            if not reply:
                return "❌ OpenClaw returned empty response"
            return reply
    except httpx.TimeoutException:
        return "⏱ OpenClaw timeout after {}s".format(timeout_sec)
    except httpx.ConnectError:
        return "❌ OpenClaw Gateway tidak aktif. Coba /restart_openclaw dulu."
    except Exception as e:
        log.error("OpenClaw error: %s", e)
        return "❌ OpenClaw error: {}".format(e)

# === AI CLIENT (multi-model) ==================================================
def get_user_model(user_id):
    """Get the model config for a user."""
    model_key = user_model_selection.get(user_id, DEFAULT_MODEL)
    return model_key, AVAILABLE_MODELS.get(model_key, AVAILABLE_MODELS[DEFAULT_MODEL])

async def ask_ai(user_id: int, user_message: str) -> str:
    add_to_history(user_id, "user", user_message)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    messages.extend(get_history(user_id))
    
    model_key, (model_name, api_url, api_key, model_id) = get_user_model(user_id)
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                api_url,
                headers={
                    "Authorization": "Bearer " + api_key,
                    "Content-Type": "application/json"
                },
                json={
                    "model": model_id,
                    "messages": messages,
                    "temperature": 0.7,
                    "max_tokens": 4096
                }
            )
            response.raise_for_status()
            data = response.json()
            reply = data["choices"][0]["message"]["content"]
            add_to_history(user_id, "assistant", reply)
            return reply
    except httpx.TimeoutException:
        return "⏱ Timeout — {} tidak merespon.".format(model_name)
    except Exception as e:
        log.error("AI API error (%s): %s", model_name, e)
        return "❌ Error ({}): {}".format(model_name, e)

# === SEND LONG MESSAGE ========================================================
async def send_long(update, text):
    if len(text) > 4000:
        chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
        for chunk in chunks:
            await update.message.reply_text(chunk)
    else:
        await update.message.reply_text(text)

# ==============================================================================
# OPENCLAW SKILL COMMANDS
# ==============================================================================
async def cmd_skills(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """List all installed OpenClaw skills."""
    output = run_shell(
        "for f in $(find /root/.openclaw/skills -maxdepth 3 -name '*.md' "
        "-not -path '*/node_modules/*' -type f 2>/dev/null); do "
        "name=$(grep '^name:' \"$f\" 2>/dev/null | head -1 | sed 's/name: *//'); "
        "desc=$(grep '^description:' \"$f\" 2>/dev/null | head -1 | sed 's/description: *//'); "
        "if [ -n \"$name\" ]; then echo \"• $name\"; echo \"  $desc\"; echo ''; fi; done"
    )
    if not output.strip() or output.strip() == "(no output)":
        output = "Tidak ada skill yang terinstall."
    
    text = "🧩 OpenClaw Skills Terinstall\n\n" + output
    text += "\n\n💡 Gunakan /skill <nama_skill> <perintah> untuk menjalankan skill."
    text += "\nContoh: /skill debug-pro cek kenapa bot lambat"
    await send_long(update, text)

async def cmd_skill(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Execute an OpenClaw skill via Gateway."""
    text = update.message.text
    if text.startswith("/skill "):
        args = text[7:].strip()
    else:
        args = ""
    
    if not args:
        await update.message.reply_text(
            "🧩 Usage: /skill <perintah>\n\n"
            "Contoh:\n"
            "  /skill run security check\n"
            "  /skill debug kenapa bot lambat\n"
            "  /skill create cron daily jam 8 pagi\n"
            "  /skill review python code\n\n"
            "Ketik /skills untuk lihat daftar skill."
        )
        return
    
    await update.message.reply_text("🧩 Menjalankan skill via OpenClaw...")
    await update.message.chat.send_action("typing")
    
    reply = await ask_openclaw(args)
    await send_long(update, "🧩 OpenClaw Response:\n\n" + reply)

async def cmd_ask_openclaw(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Ask OpenClaw AI directly."""
    text = update.message.text
    if text.startswith("/oc "):
        query = text[4:].strip()
    elif text.startswith("/openclaw "):
        query = text[10:].strip()
    else:
        query = ""
    
    if not query:
        await update.message.reply_text("Usage: /oc <pertanyaan>\nContoh: /oc analisa security vps saya")
        return
    
    await update.message.reply_text("⚡ Asking OpenClaw AI...")
    await update.message.chat.send_action("typing")
    
    reply = await ask_openclaw(query)
    await send_long(update, reply)

# ==============================================================================
# MODEL SWITCHING
# ==============================================================================
async def cmd_model(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Switch AI model. Usage: /model <name>"""
    user_id = update.effective_user.id
    text = update.message.text
    target = text[7:].strip().lower() if text.startswith("/model ") else ""
    
    current_key, (current_name, _, _, _) = get_user_model(user_id)
    
    if not target or target == "list":
        lines = ["🤖 Model AI Tersedia\n"]
        for key, (name, url, api_key, model_id) in AVAILABLE_MODELS.items():
            marker = " ✅" if key == current_key else ""
            lines.append("  /model {} — {}{}".format(key, name, marker))
        lines.append("\n📌 Saat ini: {} ({})".format(current_name, current_key))
        lines.append("\nContoh: /model kimi")
        await update.message.reply_text("\n".join(lines))
        return
    
    if target not in AVAILABLE_MODELS:
        await update.message.reply_text(
            "❌ Model '{}' tidak tersedia.\n"
            "Model tersedia: {}".format(target, ", ".join(AVAILABLE_MODELS.keys()))
        )
        return
    
    target_info = AVAILABLE_MODELS[target]
    user_model_selection[user_id] = target
    
    # Reset conversation on model switch
    conversation_history.pop(user_id, None)
    
    await update.message.reply_text(
        "✅ Model diganti!\n\n"
        "🔄 {} → {}\n"
        "🆔 {}\n\n"
        "💬 History direset. Mulai chat!".format(
            current_name, target_info[0], target_info[3]
        )
    )

# ==============================================================================
# CHAT HANDLERS
# ==============================================================================
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /start — delegates to SkillMenuEngine for interactive menu."""
    # Fallback: if SkillMenuEngine is not available, show text-only
    user_id = update.effective_user.id
    _, (model_name, _, _, _) = get_user_model(user_id)
    text = (
        "👋 Halo! Saya AI Assistant.\n\n"
        "🤖 Model: {} (ketik /model untuk ganti)\n\n"
        "💬 Chat biasa = dijawab AI\n"
        "🧩 /skill <perintah> = jalankan OpenClaw skill\n"
        "⚡ /oc <pertanyaan> = tanya OpenClaw AI\n\n"
        "📌 /help — Bantuan lengkap\n"
        "/model — Ganti model AI\n"
        "/skills — Lihat semua skill\n"
        "/reset — Reset percakapan"
    ).format(model_name)
    if is_admin(user_id):
        text += "\n\n🔧 Admin: /debug"
    await update.message.reply_text(text)

async def cmd_reset(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    conversation_history.pop(user_id, None)
    await update.message.reply_text("🔄 Percakapan direset!")

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    _, (model_name, _, _, _) = get_user_model(user_id)
    text = (
        "🤖 Bot 2 — Help\n\n"
        "💬 Chat (Model: {}):\n"
        "  Pesan biasa = Dijawab AI\n"
        "  /model = Ganti model AI\n"
        "  /reset = Reset history\n\n"
        "🧩 OpenClaw Skills:\n"
        "  /skills — Semua skill\n"
        "  /skill <perintah> — Jalankan skill\n"
        "  /oc <pertanyaan> — Tanya OpenClaw\n\n"
        "🛠 Utility:\n"
        "  /note /notes /delnote — Catatan\n"
        "  /renamenote <no> <judul> — Ganti judul\n"
        "  /remind <waktu> <pesan> — Pengingat\n"
        "  /calc <expr> — Kalkulator\n"
        "  /search <query> — Cari web (DuckDuckGo)\n"
        "  /threads — List percakapan aktif\n"
        "  /trend — AI funnel trend research\n"
        "  /trend_full — Summary + JSON file\n"
        "  /trend_window — Fixed window research\n"
        "  /trend_help — Trend commands help\n"
        "  📸 Kirim foto — OCR\n"
    ).format(model_name)
    if is_admin(user_id):
        text += (
            "\n🔧 Debug (Admin):\n"
            "  /debug — Panel debug\n"
            "  /status — VPS info\n"
            "  /logs /logs1 /logclaw — Logs\n"
            "  /services — Status services\n"
            "  /restart_bot1 /restart_bot2\n"
            "  /restart_openclaw\n"
            "  /ps /df /ip /ports /netstat\n"
            "  /sh <cmd> — Shell\n"
            "  /ping <host> — Ping\n"
            "  /security — Audit\n"
            "  /find <nama> — Cari file\n"
            "  /backup — Backup\n"
            "  /download <url> — Download\n"
            "  /botinfo — Bot config\n"
        )
    await update.message.reply_text(text)

# ==============================================================================
# SKILL: OCR (photo handler) + auto-save to notes
# ==============================================================================
async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """OCR photo. If caption starts with 'ocr', auto-save to notes."""
    if not update.message or not update.message.photo:
        return
    
    caption = (update.message.caption or "").strip()
    save_title = None
    if caption.lower().startswith("ocr"):
        save_title = caption[3:].strip() or ("OCR " + datetime.now().strftime("%d/%m %H:%M"))
    
    await update.message.reply_text("📸 Membaca teks dari gambar...")
    output = await ocr_image(ctx, update.message.photo[-1])
    
    if not output:
        await update.message.reply_text("❌ Tidak dapat membaca teks / OCR gagal.")
        return
    
    if save_title:
        notes = load_notes()
        note = {
            "title": save_title,
            "text": output,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "user": update.effective_user.id,
            "source": "ocr"
        }
        notes.append(note)
        save_notes(notes)
        await update.message.reply_text("📸 OCR + 📝 Disimpan!\n📌 {}\n\nKetik /viewnote {} untuk cek.".format(save_title, len(notes)))
    else:
        await update.message.reply_text("📝 Hasil OCR:\n\n" + output + "\n\n💡 Tip: Reply '/save Judul' untuk simpan.")

# ==============================================================================
# SKILL: FILE IMPORT (CSV/XLSX/TXT → auto-save to notes)
# ==============================================================================
async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle uploaded files. CSV/XLSX/TXT are read and saved to notes."""
    if not update.message or not update.message.document:
        return
    
    doc = update.message.document
    filename = doc.file_name or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    supported = {"csv", "xlsx", "xls", "txt", "md", "json"}
    if ext not in supported:
        return
    
    caption = (update.message.caption or "").strip()
    title = caption if caption else filename.rsplit(".", 1)[0].replace("_", " ").replace("-", " ")
    
    await update.message.reply_text("📄 Membaca & menyimpan file {}...".format(filename))
    
    content, fname, _ = await read_document_content(ctx, doc)
    
    if not content:
        await update.message.reply_text("❌ File kosong atau gagal dibaca.")
        return
    
    if len(content) > 50000:
        content = content[:50000] + "\n\n... (truncated)"
    
    notes = load_notes()
    note = {
        "title": title,
        "text": content,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "user": update.effective_user.id,
        "source": "file:" + filename
    }
    notes.append(note)
    save_notes(notes)
    
    line_count = len([l for l in content.split("\n") if l.strip()])
    preview = content[:200] + "..."
    await update.message.reply_text(
        "📄 File disimpan ke Notes!\n\n"
        "📌 {}\n"
        "📊 {} baris\n\n"
        "Preview:\n{}\n\n"
        "💡 Tip: Reply '/save Judul Lain' jika ingin simpan ulang dengan nama beda.".format(title, line_count, preview)
    )

# ==============================================================================
# SKILL: URL CHECKER
# ==============================================================================
async def cmd_curl(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    url = text[6:].strip() if text.startswith("/curl ") else ""
    if not url:
        await update.message.reply_text("Usage: /curl <url>\nContoh: /curl https://google.com")
        return
    if not url.startswith("http"):
        url = "https://" + url
    await update.message.reply_text("🌐 Checking " + url + " ...")
    output = run_shell(
        "curl -sI -o /dev/null -w 'Status: %{{http_code}}\\nTime: %{{time_total}}s\\nIP: %{{remote_ip}}' '{}' 2>&1".format(url),
        timeout=15
    )
    headers = run_shell("curl -sI '{}' 2>&1 | head -10".format(url), timeout=10)
    await update.message.reply_text("🌐 {}\n\n{}\n\n📋 Headers:\n{}".format(url, output, headers))

# ==============================================================================
# HELPERS (OCR & FILE READ)
# ==============================================================================
async def ocr_image(ctx, photo):
    """Run OCR on a photo object."""
    try:
        file = await ctx.bot.get_file(photo.file_id)
        tmp_path = "/tmp/ocr_{}.jpg".format(int(time.time()))
        await file.download_to_drive(tmp_path)
        output = run_shell("python3 /root/.openclaw/scripts/ocr_image.py '{}' 2>&1".format(tmp_path))
        run_shell("rm -f '{}'".format(tmp_path))
        return output.strip() if output.strip() and output.strip() != "(no output)" else None
    except Exception as e:
        log.error("OCR helper error: %s", e)
        return None

async def read_document_content(ctx, doc):
    """Read content from a document object."""
    filename = doc.file_name or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        file = await ctx.bot.get_file(doc.file_id)
        tmp_path = "/tmp/upload_{}_{}".format(int(time.time()), filename)
        await file.download_to_drive(tmp_path)
        
        content = ""
        if ext in ["txt", "md"]:
            with open(tmp_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        elif ext == "json":
            with open(tmp_path, "r", encoding="utf-8", errors="replace") as f:
                content = json.dumps(json.load(f), indent=2, ensure_ascii=False)
        elif ext == "csv":
            rows = []
            with open(tmp_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i == 0: rows.append("[HEADER] " + " | ".join(row))
                    else: rows.append("[{}] {}".format(i, " | ".join(row)))
            content = "\n".join(rows)
        elif ext in ["xlsx", "xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_rows = ["=== Sheet: {} ===".format(sheet_name)]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    cells = [str(c) if c is not None else "" for c in row]
                    if i == 0: sheet_rows.append("[HEADER] " + " | ".join(cells))
                    else: sheet_rows.append("[{}] {}".format(i, " | ".join(cells)))
                sheets.append("\n".join(sheet_rows))
            wb.close()
            content = "\n\n".join(sheets)
        
        run_shell("rm -f '{}'".format(tmp_path))
        return content, filename, ext
    except Exception as e:
        log.error("Read doc error: %s", e)
        return None, filename, ext

# ==============================================================================
# SKILL: NOTES / PROMPT MANAGER
# ==============================================================================
async def cmd_note(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Save a note or prompt. Use | to separate title from content.
    /note My Prompt Title | actual prompt content here...
    /note Just a short note without title
    """
    content = update.message.text[6:].strip() if update.message.text.startswith("/note ") else ""
    if not content:
        await update.message.reply_text(
            "📝 Simpan catatan / prompt\n\n"
            "Format:\n"
            "  /note Judul | isi prompt panjang...\n"
            "  /note catatan singkat\n\n"
            "Contoh:\n"
            "  /note SEO Writer | Kamu adalah SEO expert. Tulis artikel...\n"
            "  /note Beli domain besok\n\n"
            "Commands:\n"
            "  /notes — Lihat daftar\n"
            "  /viewnote 3 — Lihat isi lengkap #3\n"
            "  /findnote keyword — Cari\n"
            "  /renamenote 3 judul baru — Ganti judul\n"
            "  /delnote 3 — Hapus #3"
        )
        return
    
    # Parse title | content format
    if "|" in content:
        parts = content.split("|", 1)
        title = parts[0].strip()
        body = parts[1].strip()
    else:
        # Short note: use first 50 chars as title, full text as body
        title = content[:50] + ("..." if len(content) > 50 else "")
        body = content
    
    notes = load_notes()
    note = {
        "title": title,
        "text": body,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "user": update.effective_user.id
    }
    notes.append(note)
    save_notes(notes)
    
    preview = body[:100] + ("..." if len(body) > 100 else "")
    await update.message.reply_text(
        "📝 Prompt #{} disimpan!\n\n"
        "📌 {}\n"
        "📄 {}\n\n"
        "Ketik /viewnote {} untuk lihat lengkap".format(
            len(notes), title, preview, len(notes)
        )
    )

async def cmd_notes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """List all saved notes/prompts (titles only)."""
    notes = load_notes()
    if not notes:
        await update.message.reply_text("📝 Belum ada catatan.\n\nKetik /note untuk cara simpan.")
        return
    
    lines = ["📝 Prompt & Catatan ({} total)\n".format(len(notes))]
    for i, n in enumerate(notes):
        title = n.get("title", n.get("text", "")[:40])
        body_len = len(n.get("text", ""))
        time_str = n.get("time", "")
        lines.append("{}. 📌 {} ({} chars, {})".format(i+1, title, body_len, time_str))
    
    lines.append("\n💡 /viewnote <no> untuk lihat isi lengkap")
    lines.append("🔍 /findnote <keyword> untuk cari")
    await send_long(update, "\n".join(lines))

async def cmd_viewnote(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """View full content of a note/prompt."""
    num = update.message.text[10:].strip() if update.message.text.startswith("/viewnote ") else ""
    if not num or not num.isdigit():
        await update.message.reply_text("Usage: /viewnote <nomor>\nContoh: /viewnote 3")
        return
    
    idx = int(num) - 1
    notes = load_notes()
    if idx < 0 or idx >= len(notes):
        await update.message.reply_text("❌ Nomor tidak valid. Ketik /notes")
        return
    
    n = notes[idx]
    title = n.get("title", "Catatan")
    body = n.get("text", "")
    time_str = n.get("time", "")
    
    text = "📝 #{} — {}\n📅 {}\n\n{}".format(num, title, time_str, body)
    await send_long(update, text)

async def cmd_findnote(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Search notes/prompts by keyword."""
    keyword = update.message.text[10:].strip() if update.message.text.startswith("/findnote ") else ""
    if not keyword:
        await update.message.reply_text("Usage: /findnote <keyword>\nContoh: /findnote SEO")
        return
    
    notes = load_notes()
    results = []
    kw_lower = keyword.lower()
    for i, n in enumerate(notes):
        title = n.get("title", "")
        body = n.get("text", "")
        if kw_lower in title.lower() or kw_lower in body.lower():
            results.append((i+1, n))
    
    if not results:
        await update.message.reply_text("🔍 Tidak ditemukan catatan dengan '{}'\n\nKetik /notes untuk lihat semua.".format(keyword))
        return
    
    lines = ["🔍 Hasil cari '{}' ({} ditemukan)\n".format(keyword, len(results))]
    for num, n in results:
        title = n.get("title", n.get("text", "")[:40])
        lines.append("{}. 📌 {}".format(num, title))
    lines.append("\n💡 /viewnote <no> untuk lihat isi")
    await send_long(update, "\n".join(lines))

async def cmd_renamenote(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Rename a note/prompt by number."""
    args = update.message.text[12:].strip() if update.message.text.startswith("/renamenote ") else ""
    parts = args.split(" ", 1)
    if len(parts) < 2 or not parts[0].isdigit():
        await update.message.reply_text("Usage: /renamenote <nomor> <judul baru>\nContoh: /renamenote 3 Mastering Prompt GPT")
        return
    
    idx = int(parts[0]) - 1
    new_title = parts[1].strip()
    
    notes = load_notes()
    if idx < 0 or idx >= len(notes):
        await update.message.reply_text("❌ Nomor tidak valid. Ketik /notes")
        return
    
    old_title = notes[idx].get("title", "Untitled")
    notes[idx]["title"] = new_title
    # Update title in text if it was part of title|content format? No, just store as title property.
    
    save_notes(notes)
    await update.message.reply_text(
        "📝 Judul diganti!\n\n"
        "❌ {}\n"
        "✅ {}".format(old_title, new_title)
    )

async def cmd_delnote(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Delete a note/prompt by number."""
    num = update.message.text[9:].strip() if update.message.text.startswith("/delnote ") else ""
    if not num or not num.isdigit():
        await update.message.reply_text("Usage: /delnote <nomor>")
        return
    idx = int(num) - 1
    notes = load_notes()
    if idx < 0 or idx >= len(notes):
        await update.message.reply_text("❌ Nomor tidak valid. Ketik /notes")
        return
    removed = notes.pop(idx)
    save_notes(notes)
    
    title = removed.get("title", removed.get("text", "")[:40])
    await update.message.reply_text("🗑 Prompt #{} dihapus: \"{}\"".format(num, title))

async def cmd_save(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Save content from a replied message."""
    if not update.message.reply_to_message:
        await update.message.reply_text("Usage: Reply pesan yg mau disimpan dengan /save <judul>")
        return
    
    args = update.message.text[6:].strip() if update.message.text.startswith("/save ") else ""
    title = args if args else "Saved " + datetime.now().strftime("%d/%m %H:%M")
    
    target = update.message.reply_to_message
    content = ""
    source = "reply"
    
    await update.message.reply_text("⏳ Menyimpan...")
    
    if target.text:
        content = target.text
        source = "reply:text"
    elif target.photo:
        content = await ocr_image(ctx, target.photo[-1])
        if not content:
            await update.message.reply_text("❌ Gagal OCR foto tersebut.")
            return
        source = "reply:ocr"
    elif target.document:
        content, fname, _ = await read_document_content(ctx, target.document)
        if not content:
            await update.message.reply_text("❌ Gagal membaca file/dokumen (format tidak support?).")
            return
        source = "reply:file:" + fname
    elif target.caption:
        content = target.caption
        source = "reply:caption"
    else:
        await update.message.reply_text("❌ Tipe pesan tidak didukung untuk disimpan.")
        return
    
    # Truncate if too long
    if len(content) > 50000:
        content = content[:50000] + "\n\n... (truncated)"
    
    notes = load_notes()
    note = {
        "title": title,
        "text": content,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "user": update.effective_user.id,
        "source": source
    }
    notes.append(note)
    save_notes(notes)
    
    preview = content[:100] + "..."
    await update.message.reply_text(
        "📝 Berhasil disimpan ke Notes!\n\n"
        "📌 {}\n"
        "📄 {} chars\n\n"
        "Ketik /viewnote {} untuk lihat lengkap".format(title, len(content), len(notes))
    )

async def cmd_exportnotes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Export the notes database (JSON) to chat."""
    await update.message.reply_text("📦 Menyiapkan file notes...")
    try:
        from pathlib import Path
        db_path = Path("/root/.bot2_notes.json")
        if not db_path.exists():
            await update.message.reply_text("❌ Database tidak ditemukan (belum ada notes?).")
            return
        
        await update.message.reply_document(document=open(db_path, "rb"), filename="bot_notes_backup.json")
    except Exception as e:
        await update.message.reply_text("❌ Gagal export: {}".format(e))

# ==============================================================================
# SKILL: REMINDER
# ==============================================================================
async def cmd_remind(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = update.message.text[8:].strip() if update.message.text.startswith("/remind ") else ""
    if not args:
        await update.message.reply_text(
            "⏰ Usage: /remind <waktu> <pesan>\n\n"
            "Contoh:\n"
            "  /remind 5m Cek email\n"
            "  /remind 1h Meeting zoom\n"
            "  /remind 1d Bayar hosting"
        )
        return
    parts = args.split(None, 1)
    if len(parts) < 2:
        await update.message.reply_text("❌ Format: /remind <waktu> <pesan>")
        return
    
    match = re.match(r'^(\d+)(m|h|d)$', parts[0].lower())
    if not match:
        await update.message.reply_text("❌ Format waktu: 5m, 1h, 1d")
        return
    
    amount = int(match.group(1))
    unit = match.group(2)
    seconds = amount * (60 if unit=="m" else 3600 if unit=="h" else 86400)
    unit_label = {"m":"menit","h":"jam","d":"hari"}[unit]
    
    if seconds > 7*86400:
        await update.message.reply_text("❌ Max 7 hari.")
        return
    
    chat_id = update.effective_chat.id
    message = parts[1]
    ctx.job_queue.run_once(
        reminder_callback, seconds,
        data={"chat_id": chat_id, "message": message},
        name="remind_{}_{}".format(chat_id, int(time.time()))
    )
    due = (datetime.now() + timedelta(seconds=seconds)).strftime("%H:%M %d/%m")
    await update.message.reply_text("⏰ Set! \"{}\" dalam {} {} (jam {})".format(message, amount, unit_label, due))

async def reminder_callback(context: ContextTypes.DEFAULT_TYPE):
    data = context.job.data
    await context.bot.send_message(chat_id=data["chat_id"], text="⏰ REMINDER!\n\n📝 " + data["message"])

# ==============================================================================
# SKILL: CALCULATOR
# ==============================================================================
async def cmd_calc(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    expr = update.message.text[6:].strip() if update.message.text.startswith("/calc ") else ""
    if not expr:
        await update.message.reply_text("🧮 Usage: /calc <ekspresi>\nContoh: /calc 25 * 4 + 10")
        return
    safe_pattern = re.compile(r'^[0-9+\-*/().%\s,sqrtpowabsceilfloorlogpie]+$', re.IGNORECASE)
    if not safe_pattern.match(expr):
        await update.message.reply_text("❌ Hanya angka dan operator matematika.")
        return
    try:
        safe_expr = expr.replace("sqrt","math.sqrt").replace("pow","math.pow")
        safe_expr = safe_expr.replace("ceil","math.ceil").replace("floor","math.floor")
        safe_expr = safe_expr.replace("log","math.log").replace("pi","math.pi")
        result = eval(safe_expr, {"__builtins__": {}, "math": math, "abs": abs}, {})
        if isinstance(result, float) and result == int(result):
            formatted = "{:,}".format(int(result))
        elif isinstance(result, float):
            formatted = "{:,.4f}".format(result)
        else:
            formatted = "{:,}".format(result)
        await update.message.reply_text("🧮 {} = {}".format(expr, formatted))
    except Exception as e:
        await update.message.reply_text("❌ Error: {}".format(e))

# ==============================================================================
# SKILL: SECURITY AUDIT
# ==============================================================================
async def cmd_security(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    await update.message.reply_text("🔒 Running security audit...")
    checks = []
    checks.append("📌 SSH Config:\n" + run_shell("grep -E '^(PermitRootLogin|PasswordAuthentication|Port)' /etc/ssh/sshd_config 2>/dev/null"))
    checks.append("📌 Open Ports:\n" + run_shell("ss -tlnp | head -15"))
    checks.append("📌 Firewall:\n" + run_shell("(iptables -L -n 2>/dev/null | head -10) || echo 'No firewall'"))
    checks.append("📌 Failed Logins:\n" + run_shell("lastb 2>/dev/null | head -5 || echo 'N/A'"))
    checks.append("📌 SUID files: " + run_shell("find / -perm -4000 -type f 2>/dev/null | wc -l").strip())
    result = "🔒 Security Audit\n\n" + "\n\n".join(checks)
    await send_long(update, result)

# ==============================================================================
# SKILL: FIND FILE
# ==============================================================================
async def cmd_find(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    query = update.message.text[6:].strip() if update.message.text.startswith("/find ") else ""
    if not query:
        await update.message.reply_text("Usage: /find <nama>")
        return
    output = run_shell("find /root -maxdepth 4 -name '*{}*' 2>/dev/null | head -30".format(query))
    await update.message.reply_text("🔍 '{}'\n\n{}".format(query, output))

# ==============================================================================
# SKILL: NETWORK MONITOR
# ==============================================================================
async def cmd_netstat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("ss -tunap | head -25")
    await update.message.reply_text("📊 Connections\n\n" + output)

async def cmd_ports(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("ss -tlnp")
    await update.message.reply_text("🔌 Open Ports\n\n" + output)

# ==============================================================================
# SKILL: BACKUP
# ==============================================================================
async def cmd_backup(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bdir = "/root/backups/" + ts
    await update.message.reply_text("📦 Creating backup...")
    output = run_shell(
        "mkdir -p '{}' && "
        "cp /root/telegram_middleware.py '{}/' 2>/dev/null; "
        "cp /root/telegram_bot2.py '{}/' 2>/dev/null; "
        "cp /root/.openclaw/openclaw.json '{}/' 2>/dev/null; "
        "cp /root/.bot2_notes.json '{}/' 2>/dev/null; "
        "echo 'Files:' && ls -la '{}' && echo '' && du -sh '{}'".format(
            bdir, bdir, bdir, bdir, bdir, bdir, bdir
        )
    )
    await update.message.reply_text("📦 Backup done!\n📁 {}\n\n{}".format(bdir, output))

# ==============================================================================
# SKILL: DOWNLOAD
# ==============================================================================
async def cmd_download(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    url = update.message.text[10:].strip() if update.message.text.startswith("/download ") else ""
    if not url:
        await update.message.reply_text("Usage: /download <url>")
        return
    if not url.startswith("http"):
        url = "https://" + url
    await update.message.reply_text("⬇️ Downloading...")
    output = run_shell("mkdir -p /root/downloads && wget -P /root/downloads '{}' 2>&1 | tail -5".format(url), timeout=60)
    await update.message.reply_text("⬇️ Result:\n\n" + output)

# ==============================================================================
# BOT DEVELOPMENT (live edit/debug via Telegram)
# ==============================================================================
BOT_FILE = "/root/telegram_bot2.py"

async def cmd_viewbot(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """View bot source code. /viewbot 10 30 = show lines 10-30."""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    args = update.message.text[9:].strip() if update.message.text.startswith("/viewbot ") else ""
    if not args:
        # Show line count + overview
        total = run_shell("wc -l < '{}'".format(BOT_FILE)).strip()
        funcs = run_shell("grep -n 'async def \\|^def ' '{}' | head -40".format(BOT_FILE))
        await send_long(update, "💻 Bot Source ({} lines)\n\n📌 Functions:\n{}\n\nUsage: /viewbot 10 30".format(total, funcs))
        return
    parts = args.split()
    try:
        start = int(parts[0])
        end = int(parts[1]) if len(parts) > 1 else start + 30
        end = min(end, start + 80)  # Max 80 lines at a time
    except ValueError:
        # Search for text instead
        keyword = args
        output = run_shell("grep -n '{}' '{}' | head -20".format(keyword, BOT_FILE))
        if output.strip() == "(no output)":
            await update.message.reply_text("🔍 '{}' tidak ditemukan.".format(keyword))
        else:
            await send_long(update, "🔍 Hasil cari '{}':\n\n{}".format(keyword, output))
        return
    output = run_shell("sed -n '{},{}p' '{}'".format(start, end, BOT_FILE))
    await send_long(update, "💻 Lines {}-{}:\n\n{}".format(start, end, output))

async def cmd_editbot(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Replace lines. /editbot 42 43 | new content here"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    args = update.message.text[9:].strip() if update.message.text.startswith("/editbot ") else ""
    if not args or "|" not in args:
        await update.message.reply_text(
            "✏️ Edit Bot Source\n\n"
            "Format: /editbot <line_start> <line_end> | new content\n\n"
            "Contoh:\n"
            "  /editbot 42 42 | AI_MODEL = 'deepseek-chat'\n"
            "  /editbot 100 105 | # kode baru di sini\n\n"
            "⚠️ Tip: pakai /viewbot dulu, lalu /checkbot setelah edit."
        )
        return
    
    meta, content = args.split("|", 1)
    parts = meta.strip().split()
    try:
        start = int(parts[0])
        end = int(parts[1]) if len(parts) > 1 else start
    except (ValueError, IndexError):
        await update.message.reply_text("❌ Format: /editbot <line_start> <line_end> | new content")
        return
    
    new_content = content.strip()
    
    # Backup first
    run_shell("cp '{}' '{}.bak'".format(BOT_FILE, BOT_FILE))
    
    # Read file, replace lines
    try:
        with open(BOT_FILE, "r") as f:
            lines = f.readlines()
        
        total = len(lines)
        if start < 1 or end > total or start > end:
            await update.message.reply_text("❌ Line range {}-{} invalid (file: {} lines)".format(start, end, total))
            return
        
        old_content = "".join(lines[start-1:end])
        
        # Replace lines
        new_lines = lines[:start-1] + [new_content + "\n"] + lines[end:]
        
        with open(BOT_FILE, "w") as f:
            f.writelines(new_lines)
        
        await update.message.reply_text(
            "✏️ Lines {}-{} diganti!\n\n"
            "❌ Old:\n{}\n"
            "✅ New:\n{}\n\n"
            "📌 /checkbot — cek syntax\n"
            "/restart_bot2 — apply".format(
                start, end, old_content.strip()[:500], new_content[:500]
            )
        )
    except Exception as e:
        await update.message.reply_text("❌ Error: {}".format(e))

async def cmd_addbot(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Insert code at a specific line. /addbot 100 | new code here"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    args = update.message.text[8:].strip() if update.message.text.startswith("/addbot ") else ""
    if not args or "|" not in args:
        await update.message.reply_text(
            "➕ Insert Code\n\n"
            "Format: /addbot <after_line> | new code\n\n"
            "Contoh: /addbot 100 | # New function\n"
            "Sisipkan kode di bawah line 100."
        )
        return
    
    meta, content = args.split("|", 1)
    try:
        after_line = int(meta.strip())
    except ValueError:
        await update.message.reply_text("❌ Format: /addbot <line_number> | code")
        return
    
    new_content = content.strip()
    run_shell("cp '{}' '{}.bak'".format(BOT_FILE, BOT_FILE))
    
    try:
        with open(BOT_FILE, "r") as f:
            lines = f.readlines()
        
        if after_line < 0 or after_line > len(lines):
            await update.message.reply_text("❌ Line {} invalid (file: {} lines)".format(after_line, len(lines)))
            return
        
        new_lines = lines[:after_line] + [new_content + "\n"] + lines[after_line:]
        
        with open(BOT_FILE, "w") as f:
            f.writelines(new_lines)
        
        await update.message.reply_text(
            "➕ Code disisipkan setelah line {}!\n\n"
            "{}\n\n"
            "📌 /checkbot — cek syntax\n"
            "/restart_bot2 — apply".format(after_line, new_content[:500])
        )
    except Exception as e:
        await update.message.reply_text("❌ Error: {}".format(e))

async def cmd_checkbot(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Check bot syntax without restarting."""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("python3 -m py_compile '{}' 2>&1".format(BOT_FILE))
    if output.strip() == "(no output)":
        total = run_shell("wc -l < '{}'".format(BOT_FILE)).strip()
        await update.message.reply_text(
            "✅ Syntax OK!\n\n"
            "💻 {} lines\n"
            "📌 /restart_bot2 untuk apply changes".format(total)
        )
    else:
        await update.message.reply_text("❌ Syntax Error:\n\n" + output)

async def cmd_undoedit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Restore from backup."""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    bak = BOT_FILE + ".bak"
    exists = run_shell("test -f '{}' && echo 'yes'".format(bak)).strip()
    if exists != "yes":
        await update.message.reply_text("❌ Tidak ada backup file.")
        return
    run_shell("cp '{}' '{}'".format(bak, BOT_FILE))
    await update.message.reply_text(
        "↩️ Restored dari backup!\n\n"
        "📌 /checkbot — cek syntax\n"
        "/restart_bot2 — apply"
    )

async def cmd_pip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Install/upgrade pip packages. /pip install openpyxl"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    args = update.message.text[5:].strip() if update.message.text.startswith("/pip ") else ""
    if not args:
        await update.message.reply_text(
            "📦 Pip Manager\n\n"
            "/pip install <package>\n"
            "/pip upgrade <package>\n"
            "/pip list\n"
            "/pip freeze"
        )
        return
    
    # Safety: block dangerous pip commands
    if any(x in args for x in ["uninstall", "--target /", "&&", ";", "|"]):
        await update.message.reply_text("⛔ Command tidak diizinkan.")
        return
    
    if args == "list":
        output = run_shell("pip3 list 2>/dev/null | head -50")
    elif args == "freeze":
        output = run_shell("pip3 freeze 2>/dev/null | head -50")
    elif args.startswith("upgrade "):
        pkg = args[8:].strip()
        await update.message.reply_text("📦 Upgrading {}...".format(pkg))
        output = run_shell("pip3 install --upgrade {} 2>&1".format(pkg), timeout=60)
    elif args.startswith("install "):
        pkg = args[8:].strip()
        await update.message.reply_text("📦 Installing {}...".format(pkg))
        output = run_shell("pip3 install {} 2>&1".format(pkg), timeout=60)
    else:
        output = run_shell("pip3 {} 2>&1".format(args), timeout=30)
    
    await send_long(update, "📦 Pip:\n\n" + output)

async def cmd_debugerr(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Show recent error logs."""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell(
        "journalctl -u telegram-bot2 --no-pager -n 100 2>/dev/null | "
        "grep -i 'error\\|traceback\\|exception\\|critical\\|failed' | "
        "tail -20"
    )
    if not output.strip() or output.strip() == "(no output)":
        await update.message.reply_text("✅ Tidak ada error terbaru! Bot berjalan normal.")
    else:
        await send_long(update, "⚠️ Recent Errors:\n\n" + output)

# ==============================================================================
# DEBUG / ADMIN HANDLERS
# ==============================================================================
async def cmd_debug(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    text = (
        "🔧 Debug Panel\n\n"
        "📊 Monitoring:\n"
        "  /status /ps /df /ip /ports /netstat\n\n"
        "📋 Logs:\n"
        "  /logs /logs1 /logclaw\n"
        "  /debugerr — Error terbaru\n\n"
        "🔄 Services:\n"
        "  /services\n"
        "  /restart_bot1 /restart_bot2 /restart_openclaw\n\n"
        "💻 Dev Tools:\n"
        "  /viewbot [line|keyword] — Lihat source\n"
        "  /editbot <L1> <L2> | code — Edit lines\n"
        "  /addbot <line> | code — Sisipkan kode\n"
        "  /checkbot — Syntax check\n"
        "  /undoedit — Undo edit terakhir\n"
        "  /pip install/list/upgrade\n\n"
        "💻 Shell:\n"
        "  /sh <cmd> /ping <host> /find <nama>\n"
        "  /download <url> /curl <url>\n\n"
        "🔒 Security:\n"
        "  /security /backup\n\n"
        "🧩 OpenClaw:\n"
        "  /skills /skill <cmd> /oc <query>\n\n"
        "🔍 /botinfo"
    )
    await update.message.reply_text(text)

async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell(
        "echo 'Uptime:' && uptime && echo '' && "
        "echo 'Memory:' && free -h && echo '' && "
        "echo 'Disk:' && df -h / && echo '' && "
        "echo 'Top CPU:' && ps aux --sort=-%cpu | head -6"
    )
    await update.message.reply_text("📊 VPS Status\n\n" + output)

async def cmd_logs(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("journalctl -u telegram-bot2 --no-pager -n 20")
    await send_long(update, "📋 Bot 2 Logs\n\n" + output)

async def cmd_logs1(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("journalctl -u telegram-bot1 --no-pager -n 20")
    await send_long(update, "📋 Bot 1 Logs\n\n" + output)

async def cmd_logclaw(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("tail -30 /root/.openclaw/bot.log 2>/dev/null || echo 'No log'")
    await send_long(update, "📋 OpenClaw Logs\n\n" + output)

async def cmd_services(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell(
        "echo 'Bot 1:' && systemctl is-active telegram-bot1 && "
        "echo 'Bot 2:' && systemctl is-active telegram-bot2 && "
        "echo 'OpenClaw:' && (pgrep -a openclaw-gateway > /dev/null && echo 'active' || echo 'inactive')"
    )
    await update.message.reply_text("🔌 Services\n\n" + output)

async def cmd_restart_bot1(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    await update.message.reply_text("🔄 Restarting Bot 1...")
    output = run_shell("systemctl restart telegram-bot1 && sleep 2 && systemctl status telegram-bot1 --no-pager | head -5")
    await update.message.reply_text(output)

async def cmd_restart_bot2(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    await update.message.reply_text("🔄 Restarting Bot 2...")
    run_shell("systemctl restart telegram-bot2")

async def cmd_restart_openclaw(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    await update.message.reply_text("🔄 Restarting OpenClaw...")
    output = run_shell(
        "pkill -f openclaw-gateway; sleep 2; "
        "nohup openclaw-gateway > /dev/null 2>&1 & sleep 3; "
        "pgrep -a openclaw-gateway || echo 'Failed'"
    )
    await update.message.reply_text(output)

async def cmd_ps(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("ps aux | grep -E 'python|node|openclaw' | grep -v grep")
    await update.message.reply_text("⚙️ Proses\n\n" + output)

async def cmd_sh(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    cmd = update.message.text[4:].strip() if update.message.text.startswith("/sh ") else ""
    if not cmd:
        await update.message.reply_text("Usage: /sh <command>")
        return
    dangerous = ["rm -rf /", "mkfs", "dd if=", "> /dev/sd"]
    for d in dangerous:
        if d in cmd:
            await update.message.reply_text("⛔ Diblokir!")
            return
    log.info("Admin shell: %s", cmd)
    await update.message.reply_text("⏳ " + cmd)
    output = run_shell(cmd, timeout=30)
    await send_long(update, output)

async def cmd_ping(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    host = update.message.text[6:].strip() if update.message.text.startswith("/ping ") else "google.com"
    if not host: host = "google.com"
    output = run_shell("ping -c 4 " + host, timeout=15)
    await update.message.reply_text("🏓 " + host + "\n\n" + output)

async def cmd_df(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("df -h")
    await update.message.reply_text("💿 Disk\n\n" + output)

async def cmd_ip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    output = run_shell("hostname -I && echo '' && curl -s ifconfig.me && echo ''")
    await update.message.reply_text("🌐 IP\n\n" + output)

async def cmd_botinfo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Akses ditolak.")
        return
    notes_count = len(load_notes())
    skill_count = run_shell("find /root/.openclaw/skills -maxdepth 3 -name '*.md' -not -path '*/node_modules/*' -type f 2>/dev/null | wc -l").strip()
    models_list = ", ".join(["{} ({})".format(v[0], k) for k, v in AVAILABLE_MODELS.items()])
    info = (
        "🤖 Bot 2 Info\n\n"
        "Default Model: " + DEFAULT_MODEL + "\n"
        "Available Models: " + models_list + "\n"
        "OpenClaw Model: " + OPENCLAW_MODEL + "\n"
        "OpenClaw Skills: " + skill_count + "\n"
        "History/user: " + str(MAX_HISTORY) + " msgs\n"
        "Active chats: " + str(len(conversation_history)) + "\n"
        "Notes saved: " + str(notes_count) + "\n"
        "Admin IDs: " + str(ADMIN_IDS)
    )
    await update.message.reply_text(info)


# === SEARCH (DuckDuckGo) ======================================================
async def cmd_search(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Search the web using DuckDuckGo."""
    if not ctx.args:
        await update.message.reply_text("Usage: /search <query>\nContoh: /search python telegram bot")
        return
    query = " ".join(ctx.args)
    await update.message.chat.send_action("typing")
    try:
        from duckduckgo_search import DDGS
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=5))
        if not results:
            await update.message.reply_text("🔍 Tidak ada hasil untuk: " + query)
            return
        text = "🔍 Hasil pencarian: {}\n\n".format(query)
        for i, r in enumerate(results, 1):
            title = r.get("title", "No title")
            href = r.get("href", "")
            body = r.get("body", "")[:150]
            text += "{}. {}\n{}\n{}\n\n".format(i, title, href, body)
        await send_long(update, text)
    except ImportError:
        await update.message.reply_text("❌ duckduckgo-search belum terinstall.\nJalankan: pip install duckduckgo-search")
    except Exception as e:
        log.error("cmd_search error: %s", e)
        await update.message.reply_text("❌ Search error: " + str(e))


# === THREADS (conversation list) ==============================================
async def cmd_threads(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """List active conversation threads."""
    if not conversation_history:
        await update.message.reply_text("📭 Belum ada thread percakapan aktif.")
        return
    text = "🧵 Active Threads\n\n"
    for uid, msgs in conversation_history.items():
        count = len(msgs)
        last_msg = msgs[-1]["content"][:60] if msgs else "-"
        text += "• User {}: {} pesan — terakhir: {}\n".format(uid, count, last_msg)
    await send_long(update, text)


# === CHAT MESSAGE HANDLER =====================================================
async def handle_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return
    user_id = update.effective_user.id
    user_text = update.message.text.strip()
    if not user_text:
        return
    log.info("User %s: %s", user_id, user_text[:100])
    await update.message.chat.send_action("typing")
    reply = await ask_ai(user_id, user_text)
    await send_long(update, reply)

# === MAIN =====================================================================
def main():
    log.info("Starting Bot 2 (Ark/GLM-4 + OpenClaw Skills + Debug)...")
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Chat
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("reset", cmd_reset))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("model", cmd_model))
    
    # OpenClaw Skills
    app.add_handler(CommandHandler("skills", cmd_skills))
    app.add_handler(CommandHandler("skill", cmd_skill))
    app.add_handler(CommandHandler("oc", cmd_ask_openclaw))
    app.add_handler(CommandHandler("openclaw", cmd_ask_openclaw))
    
    # Utility Skills
    app.add_handler(CommandHandler("note", cmd_note))
    app.add_handler(CommandHandler("notes", cmd_notes))
    app.add_handler(CommandHandler("viewnote", cmd_viewnote))
    app.add_handler(CommandHandler("findnote", cmd_findnote))
    app.add_handler(CommandHandler("renamenote", cmd_renamenote))
    app.add_handler(CommandHandler("save", cmd_save))
    app.add_handler(CommandHandler("exportnotes", cmd_exportnotes))
    app.add_handler(CommandHandler("delnote", cmd_delnote))
    app.add_handler(CommandHandler("remind", cmd_remind))
    app.add_handler(CommandHandler("calc", cmd_calc))
    app.add_handler(CommandHandler("curl", cmd_curl))
    app.add_handler(CommandHandler("search", cmd_search))
    app.add_handler(CommandHandler("threads", cmd_threads))

    # Trend Research (TrendBotRouter — PRD-03)
    trend_router = TrendBotRouter(
        youtube_api_key=YOUTUBE_API_KEY,
        available_models=AVAILABLE_MODELS,
        user_model_selection=user_model_selection,
        default_model=DEFAULT_MODEL,
        admin_ids=ADMIN_IDS,
    )
    trend_router.register_handlers(app)
    
    # Admin: Debug
    app.add_handler(CommandHandler("debug", cmd_debug))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("logs", cmd_logs))
    app.add_handler(CommandHandler("logs1", cmd_logs1))
    app.add_handler(CommandHandler("logclaw", cmd_logclaw))
    app.add_handler(CommandHandler("services", cmd_services))
    app.add_handler(CommandHandler("restart_bot1", cmd_restart_bot1))
    app.add_handler(CommandHandler("restart_bot2", cmd_restart_bot2))
    app.add_handler(CommandHandler("restart_openclaw", cmd_restart_openclaw))
    app.add_handler(CommandHandler("ps", cmd_ps))
    app.add_handler(CommandHandler("sh", cmd_sh))
    app.add_handler(CommandHandler("ping", cmd_ping))
    app.add_handler(CommandHandler("df", cmd_df))
    app.add_handler(CommandHandler("ip", cmd_ip))
    app.add_handler(CommandHandler("botinfo", cmd_botinfo))
    
    # Admin: Extra skills
    app.add_handler(CommandHandler("security", cmd_security))
    app.add_handler(CommandHandler("find", cmd_find))
    app.add_handler(CommandHandler("netstat", cmd_netstat))
    app.add_handler(CommandHandler("ports", cmd_ports))
    app.add_handler(CommandHandler("backup", cmd_backup))
    app.add_handler(CommandHandler("download", cmd_download))
    
    # Admin: Bot Development
    app.add_handler(CommandHandler("viewbot", cmd_viewbot))
    app.add_handler(CommandHandler("editbot", cmd_editbot))
    app.add_handler(CommandHandler("addbot", cmd_addbot))
    app.add_handler(CommandHandler("checkbot", cmd_checkbot))
    app.add_handler(CommandHandler("undoedit", cmd_undoedit))
    app.add_handler(CommandHandler("pip", cmd_pip))
    app.add_handler(CommandHandler("debugerr", cmd_debugerr))
    
    # Photo -> OCR (+save)
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    
    # Document -> auto-save to notes
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # Regular messages -> AI
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Dynamic Skill Menu Engine (PRD-05)
    skill_engine = SkillMenuEngine({
        "admin_ids": ADMIN_IDS,
        "ask_openclaw_fn": ask_openclaw,
        "send_long_fn": send_long,
    })
    skill_engine.register_handlers(app)
    
    log.info("Bot 2 ready with OpenClaw + 10 skills + Dynamic Menu")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
